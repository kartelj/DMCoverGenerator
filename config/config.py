from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

class Config():
    now=datetime.datetime.now()
    csvName = 'Kvartalni_TV_'+str('{:02d}_'.format(now.day))+str('{:02d}_'.format(now.month))+str(now.year)+'.xlsx'
    PATH = "baza_kvartalni.xlsx"
    NEW_FILE_PATH = csvName

    wb = Workbook()

    ws = wb.active
    ws.title = 'COVER'
    ws2 = wb.create_sheet('COVER Marko')
    ws3 = wb.create_sheet('COVER Cas')
    ws4 = wb.create_sheet('COVER Cas Marko')

    file = load_workbook(filename = PATH)

    grp_baza = file['BAZA Eq GRP i DUR']
    e2_baza = file['BAZA E2']
    shr_baza = file['BAZA Shr TV']
    cas_baza = file['BAZA CAS']

    # channel_map = {
    #     'RTS': ('RTS 1', 'RTS 2'),
    #     'B92': ('B92',),
    #     'PINK': ('PINK',),
    #     'PINK Cable': ('Pink cable total',),
    #     'Ostali': ('Other', 'RTS 3', 'RTS Drama', 'RTS Trezor', 'RTS Zivot', 'RTS Muzika', 'RTS Kolo', 'RTS Poletarac', 'RTS Nauka', 'RTS Klasika', 'RTV 1', 'RTV 2', 'Melos Kraljevo', 'Minimax', 'Agro TV', 'Kitchen TV', 'SUPERSTAR 2', 'SUPERSTAR TV', 'FILM KLUB', 'ArenaSport 1', 'Arena Sport 2', 'Arena Sport 3', 'Arena Sport 4', 'Arena Sport 5', 'Toxic TV', 'Dox TV', 'Klasik TV', 'Dexy TV', 'K1', 'Kazbuka', 'TV Doktor', 'Arena Fight', 'Balkan Trip', 'Kurir TV', 'Studio B'),
    #     'PRVA': ('PRVA',),
    #     'PRVA Cable': ('PRVA Plus', 'PRVA Max', 'PRVA World', 'PRVA Kick', 'PRVA Life', 'PRVA Files'),
    #     'Cas Media': ('N1', 'Grand Kanal', 'Cinemania', 'Nova S', 'IDJ', 'SK 1', 'SK 2', 'SK 3', 'Pikaboo', 'Vavoom', 'CineStar', 'CineStar Action', 'Discovery', 'Discovery ID', 'HGTV', 'TLC', 'DIVA - Universal', 'AXN', 'Nickelodeon', 'TV 1000', 'NickJr', 'Nova Sport', 'Animal Planet'),
    #     'Fox': ('Fox', 'Fox Crime', 'Fox Life', 'Fox Movies', 'National Geographic', '24 Kitchen'),
    #     'Happy': ('Happy',)
    # }

    channel_map = {
        'RTS': ('RTS zajedno (RTS 1+RTS 2)',),
        'B92': ('B92',),
        'PINK': ('PINK',),
        'PINK Cable': ('Pink cable total',),
        'Ostali': ('Other Total',),
        'PRVA': ('PRVA',),
        'PRVA Cable': ('Prva cable total',),
        'Cas Media': ('Cas cable total',),
        'Fox': ('Fox cable Total',),
        'Happy': ('Happy',)
    }

    cas_others_dm_pool = ['CAS Others (w/o SK, Nova Sport, Brainz)','Grand Kanal','Cinemania','IDJ','Pikaboo','Vavoom','CineStar','CineStar Action','Discovery','Discovery ID','HGTV','TLC','DIVA - Universal','AXN','Nickelodeon','TV 1000','Animal Planet','NickJr']

    cas_non_agb = ['CAS "non AGB" (SK, Nova Sport, Brainz)','Brainz','Nova Sport','Sport klub SRB']

    cas_others = ['CAS Others (w/o SK, Nova Sport, Brainz)','Grand Kanal','Cinemania','IDJ','SK 1','SK 2','SK 3','Pikaboo','Vavoom','CineStar','CineStar Action','Discovery','Discovery ID','HGTV','TLC','DIVA - Universal','AXN','Nickelodeon','TV 1000','NickJr','Nova Sport','Animal Planet']

    cas_dm_pool_channels = ['N1','Nova S'] + cas_others_dm_pool

    cas_other_channels = ['N1','Nova S'] + cas_others

    dm_pool_shr_channels = cas_dm_pool_channels + cas_non_agb

    def get_collection(self,sheet,columns):
        topics = []
        for i, cellObj in enumerate(sheet, 2):
                tup = ()
                for col in columns:
                    val = sheet.cell(row=i,column=col).value
                    if val != None:
                        if type(val) is datetime.time:
                            val = val.isoformat()
                        tup = tup + (val,)
                if len(tup) != 0:
                    topics.append(tup)
        return topics

    # Return List type []
    def column_values(self,sheet,column):
        topics = []
        for i, cellObj in enumerate(sheet, 2):
            val = sheet.cell(row=i,column=column).value
            if val != None:
                topics.append(val)
        return sorted(list(set(topics)))


    def get_collection_start_from_3(self,sheet,columns):
        topics = []
        for i, cellObj in enumerate(sheet, 3):
                tup = ()
                for col in columns:
                    val = sheet.cell(row=i,column=col).value
                    if val != None:
                        tup = tup + (val,)
                if len(tup) != 0:
                    topics.append(tup)
        return topics

    def customers(self):
        return self.column_values(self.grp_baza,20)

    def channels(self):
        return self.column_values(self.e2_baza,30)

    def e2_collection(self):
        return self.get_collection(self.e2_baza,[28,3,20,30])

    def cas_collection(self):
        return self.get_collection_start_from_3(self.cas_baza,[3,4,5,2])

    def e2_channels(self):
        e2_channels = self.channels().copy()
        e2_channels.remove('Cas Media')
        return e2_channels

    def cas_channels(self):
        return ['Cas Media', 'CAS (SK, Nova Sport, Brainz)']

    def sk_nova(self):
        return ('Sport klub SRB', 'Nova Sport', 'Brainz')

    def collection(self,db,range):
        return self.get_collection(db,range)


