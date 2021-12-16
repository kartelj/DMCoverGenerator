# -*- coding: utf-8 -*- 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter


customers = ['DM Pool','MEDIA House','MEDIA Pool','MEDIACOM','ONE MEDIA','OSTALO (A, N, O)']
cas_channels = ['Cas Media', 'CAS (SK, Nova Sport, Brainz)']
bold = Font(bold=True)
thick = Border(top=Side(border_style="thick", color="4D4D4D"),left=Side(border_style="thick", color="4D4D4D"),bottom=Side(border_style="thick", color="4D4D4D"),right=Side(border_style="thick", color="4D4D4D"))

center_align = Alignment(horizontal='center',vertical='center')
left_align = Alignment(horizontal='general',vertical='top')


def number_to_char(num):
    d = int(num/27)
    if d == 1:
        return 'A{0}'.format(chr(num+96-26).upper())
    elif d == 2:
        return 'B{0}'.format(chr(num+96-52).upper())
    else:
        return chr(num+96).upper()

def char_to_number(character):
    d = len(character)
    if d == 2:
        return ord(character[-1].lower()) - 96 + 26
    else:
        return ord(character.lower()) - 96

def next_char(letter,range):
    s = char_to_number(letter) + range
    return number_to_char(s)
      
def ColoredBorders(color, top=True, right=True, bottom=True, left=True):
    return Border(
        top=Side(style='thin', color=color) if top else None,
        right=Side(style='thin', color=color) if right else None,
        bottom=Side(style='thin', color=color) if bottom else None,
        left=Side(style='thin', color=color) if left else None,
    )

def pattern_fill(color):
    return PatternFill("solid", fgColor=color)


def row_bold_fill(ws,range,color):
    for r in ws[range]:
      for row in r:
          row.font = bold
          row.fill = PatternFill("solid", fgColor=color)

def row_fill(ws,range,color):
    for r in ws[range]:
      for row in r:
          row.fill = PatternFill("solid", fgColor=color)


def set_left_align(ws,range):
    for r in ws[range]:
        for row in r:
            row.alignment = left_align

def set_font_color(ws,range,color):
    for r in ws[range]:
        for row in r:
            row.font = Font(color=color)
            row.font = bold


def table_border(ws,min_row,max_row,min_col,max_col,number_format=None):
    for rows in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
      if rows[0].value in customers or rows[0].value == 'Total':
          border_type = thick
          f = True
      else:
          border_type = ColoredBorders(color='4D4D4D')
          f = False
      for cell in rows:
          cell.border = border_type
          cell.alignment = center_align
          if f:
              cell.font = bold
          if number_format != None and cell.value != None:
              try:
                  int(cell.value)
                  cell.number_format = number_format
              except:
                  if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0.00%'
                        except:
                            pass
                  else:
                      pass

def cas_table_border(ws,min_row,max_row,min_col,max_col,number_format=None):
    for rows in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        red = False
        if rows[0].value in customers or rows[0].value == 'Total':
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value in ['CAS Others (w/o SK, Nova Sport, Brainz)','CAS "non AGB" (SK, Nova Sport, Brainz)']:
            red = True
        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if red:
                cell.font = Font(color='FF4149')
            if number_format != None and cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = number_format
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0.00%'
                        except:
                            pass
                    else:
                        pass


def get_column_letters_for_value(ws,min_row,max_row,min_col,max_col,values):
    letters = []
    for rows in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in rows:
            for val in values:
                if str(cell.value).find(val) != -1:
                    letters.append(cell.column_letter)
    return letters


def set_cell_width(ws):
    for i in range(2, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].width = 10


def cover_style(ws,row_size):
    set_cell_width(ws)
    ws.column_dimensions['A'].width = 25

    start = row_size['total'][0]
    end = row_size['total'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    table_border(ws,start,end,1,19,'#,##0')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))
    
    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')

    start = row_size['sov_total'][0]
    end = row_size['sov_total'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = PatternFill("solid", fgColor="FDFF60")
    ws['A{0}'.format(start)].font = bold

    ws['A{0}'.format(start+1)].fill = PatternFill("solid", fgColor="FDFF60")
    ws['A{0}'.format(start+1)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    table_border(ws,start,end,1,19,'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')

    start = row_size['shr_sov'][0]
    end = row_size['shr_sov'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws.merge_cells('A{0}:S{0}'.format(start+1))

    ws['A{0}'.format(start)].fill = pattern_fill("FF4149")
    ws['A{0}'.format(start)].font = bold
    ws['A{0}'.format(start+1)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start+1)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')

    table_border(ws,start,end,1,19,'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))

    row_fill(ws,'F{0}:F{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+2,end),'BFBFBF')


    start = row_size['18-50'][0]
    end = row_size['18-50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    table_border(ws,start,end,1,19,'#,##0')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['sov_18-50'][0]
    end = row_size['sov_18-50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start)].font = bold

    ws['A{0}'.format(start+1)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start+1)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    table_border(ws,start,end,1,19,'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['shr_18-50'][0]
    end = row_size['shr_18-50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws.merge_cells('A{0}:S{0}'.format(start+1))

    ws['A{0}'.format(start)].fill = pattern_fill("FF4149")
    ws['A{0}'.format(start)].font = bold
    ws['A{0}'.format(start+1)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start+1)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')

    table_border(ws,start,end,1,19,'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))

    row_fill(ws,'F{0}:F{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+2,end),'BFBFBF')


    start = row_size['kontrolni_duration'][0]
    end = row_size['kontrolni_duration'][1]

    ws.merge_cells('A{0}:R{0}'.format(start))
    ws.merge_cells('A{0}:R{0}'.format(start+1))

    ws['A{0}'.format(start)].fill = pattern_fill("FC000B")
    ws['A{0}'.format(start)].font = bold
    ws['A{0}'.format(start+1)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start+1)].font = bold

    row_bold_fill(ws,'A{0}:R{0}'.format(start+2),'DEEAF6')

    table_border(ws,start,end,1,18,'#,##0')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))

    row_fill(ws,'E{0}:E{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'I{0}:I{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'M{0}:M{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'Q{0}:Q{1}'.format(start+1,end),'BFBFBF')


    start = row_size['duration'][0]
    end = row_size['duration'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))

    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    table_border(ws,start,end,1,19,'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')

    set_font_color(ws,'B{0}:B{1}'.format(start+2,end),'FF4149')


    start = row_size['soi'][0]
    end = row_size['soi'][1]

    ws.merge_cells('B{0}:AK{0}'.format(start))

    ws['B{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['B{0}'.format(start)].font = bold
    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    ws.merge_cells('B{0}:C{0}'.format(start+1))
    ws.merge_cells('D{0}:E{0}'.format(start+1))
    ws.merge_cells('F{0}:G{0}'.format(start+1))
    ws.merge_cells('H{0}:I{0}'.format(start+1))
    ws.merge_cells('J{0}:K{0}'.format(start+1))
    ws.merge_cells('L{0}:M{0}'.format(start+1))
    ws.merge_cells('N{0}:O{0}'.format(start+1))
    ws.merge_cells('P{0}:Q{0}'.format(start+1))
    ws.merge_cells('R{0}:S{0}'.format(start+1))
    ws.merge_cells('T{0}:U{0}'.format(start+1))
    ws.merge_cells('V{0}:W{0}'.format(start+1))
    ws.merge_cells('X{0}:Y{0}'.format(start+1))
    ws.merge_cells('Z{0}:AA{0}'.format(start+1))
    ws.merge_cells('AB{0}:AC{0}'.format(start+1))
    ws.merge_cells('AD{0}:AE{0}'.format(start+1))
    ws.merge_cells('AF{0}:AG{0}'.format(start+1))
    ws.merge_cells('AH{0}:AI{0}'.format(start+1))
    ws.merge_cells('AJ{0}:AK{0}'.format(start+1))

    row_bold_fill(ws,'A{0}:AK{0}'.format(start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=37):
        fill = False
        font_color = False
        if rows[0].value == 'DM Pool':
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value in cas_channels:
            fill = True
        if rows[0].value == 'CAS (SK, Nova Sport, Brainz)':
            font_color = True

        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if fill:
                cell.fill = pattern_fill("D8D8D8")
            if font_color:
                cell.font = Font(color='FF4149')
            if cell.value != None:
              try:
                  int(cell.value)
                  cell.number_format = '#,##0'
              except:
                  if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0.0%'
                        except:
                            pass
                  else: 
                      pass

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'K{0}:K{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'S{0}:S{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'Z{0}:Z{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'AA{0}:AA{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'AH{0}:AH{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'AI{0}:AI{1}'.format(start+1,end),'BFBFBF')


    start = row_size['cpp_4+'][0]
    end = row_size['cpp_4+'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    table_border(ws,start,end,1,19,'0.00')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['cpp_18_50'][0]
    end = row_size['cpp_18_50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    table_border(ws,start,end,1,19,'0.00')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['ratio'][0]
    end = row_size['ratio'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=19):
      if rows[0].value in customers or rows[0].value == 'Total':
          border_type = thick
          f = True
      else:
          border_type = ColoredBorders(color='4D4D4D')
          f = False
      for cell in rows:
          cell.border = border_type
          cell.alignment = center_align
          if f:
              cell.font = bold
          if cell.value != None:
              try:
                  int(cell.value)
                  cell.number_format = '0'
              except:
                  if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0%'
                        except:
                            pass
                  else:
                      pass

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')

    start = row_size['ratio_18_50'][0]
    end = row_size['ratio_18_50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=19):
      if rows[0].value in customers or rows[0].value == 'Total':
          border_type = thick
          f = True
      else:
          border_type = ColoredBorders(color='4D4D4D')
          f = False
      for cell in rows:
          cell.border = border_type
          cell.alignment = center_align
          if f:
              cell.font = bold
          if cell.value != None:
              try:
                  int(cell.value)
                  cell.number_format = '0'
              except:
                  if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0%'
                        except:
                            pass
                  else:
                      pass

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


def marko_cover_style(ws,quarter,row_size):
    set_cell_width(ws)
    ws.column_dimensions['A'].width = 25

    images = {}
    start = row_size['sov1'][0]
    end = row_size['sov1'][1]
    step = quarter+1

    images["sov_kupcima_dm_pool_others.png"] = "A{0}:{1}{2}".format(start,next_char('b',4*step+3),end)
    ws.merge_cells('A{0}:A{1}'.format(start,start+1))

    ws.merge_cells('B{0}:{1}{2}'.format(start,next_char('b',step),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',step+1),next_char('b',2*step+1),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',2*step+2),next_char('b',3*step+2),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',3*step+3),next_char('b',4*step+3),start))

    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'B{0}:{1}1'.format(start,next_char('b',4*step+3)),'DEEAF6')
    row_bold_fill(ws,'B{0}:{1}2'.format(start+1,next_char('b',4*step+3)),'DEEAF6')

    table_border(ws,start,end,1,char_to_number(next_char('b',4*step+3)),'#,##0')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))

    cols = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',4*step+3)),['FY','H1'])
    for w in cols:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+1,end),'BFBFBF')


    start = row_size['sov2'][0]
    end = row_size['sov2'][1]
    images["sov_kupcima_media_house_media_pool.png"] = "A{0}:{1}{2}".format(start,next_char('b',4*step+3),end)

    ws.merge_cells('A{0}:A{1}'.format(start,start+1))

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',step+1),next_char('b',2*step+1),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',2*step+2),next_char('b',3*step+2),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',3*step+3),next_char('b',4*step+3),start))

    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'B{1}:{0}{1}'.format(next_char('b',4*step+3),start),'DEEAF6')
    row_bold_fill(ws,'B{1}:{0}{1}'.format(next_char('b',4*step+3),start+1),'DEEAF6')

    table_border(ws,start,end,1,char_to_number(next_char('b',4*step+3)),'#,##0')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))

    cols = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',4*step+3)),['FY','H1'])
    for w in cols:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+1,end),'BFBFBF')


    start = row_size['sov3'][0]
    end = row_size['sov3'][1]
    images["sov_kupcima_mediacom_one_media.png"] = "A{0}:{1}{2}".format(start,next_char('b',4*step+3),end)

    ws.merge_cells('A{0}:A{1}'.format(start,start+1))

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',step+1),next_char('b',2*step+1),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',2*step+2),next_char('b',3*step+2),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',3*step+3),next_char('b',4*step+3),start))

    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'B{1}:{0}{1}'.format(next_char('b',4*step+3),start),'DEEAF6')
    row_bold_fill(ws,'B{1}:{0}{1}'.format(next_char('b',4*step+3),start+1),'DEEAF6')

    table_border(ws,start,end,1,char_to_number(next_char('b',4*step+3)),'#,##0')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))

    cols = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',4*step+3)),['FY','H1'])
    for w in cols:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+1,end),'BFBFBF')


    start = row_size['shr1'][0]
    end = row_size['shr1'][1]
    step = quarter + 4
    images["duration_dm_mh_mp.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)
    ws.merge_cells('A{0}:A{1}'.format(start,start+1))

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start),'DEEAF6')
    row_bold_fill(ws,'B{1}:{0}{1}'.format(next_char('b',step),start+1),'DEEAF6')

    table_border(ws,start,end,1,char_to_number(next_char('b',step)),'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))

    set_font_color(ws,'B{0}:B{1}'.format(start+2,end),'FF4149')

    cols1 = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',step)),['FY','H1'])
    for w in cols1:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+1,end),'A4A2A2')

    cols2 = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',step)),['Q'])
    for w in cols2:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+1,end),'BFBFBF')


    start = row_size['shr2'][0]
    end = row_size['shr2'][1]
    images["duration_mediacom_one_m_others.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    ws.merge_cells('A{0}:A{1}'.format(start,start+1))
    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start),'DEEAF6')
    row_bold_fill(ws,'B{1}:{0}{1}'.format(next_char('b',step),start+1),'DEEAF6')

    table_border(ws,start,end,1,char_to_number(next_char('b',step)),'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))

    set_font_color(ws,'B{0}:B{1}'.format(start+2,end),'FF4149')

    cols1 = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',step)),['FY','H1'])
    for w in cols1:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+1,end),'A4A2A2')

    cols2 = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',step)),['Q'])
    for w in cols2:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+1,end),'BFBFBF')


    start = row_size['sh'][0]
    end = row_size['sh'][1]
    images["audience_share_4plus.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('C{1}:{0}{1}'.format(next_char('b',step),start+1))

    table_border(ws,start,end,1,char_to_number(next_char('b',step)),'0')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))

    cols1 = get_column_letters_for_value(ws,start+2,end,1,char_to_number(next_char('b',step)),['FY','H1'])
    for w in cols1:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+2,end),'A4A2A2')

    cols2 = get_column_letters_for_value(ws,start+2,end,1,char_to_number(next_char('b',step)),['Q'])
    for w in cols2:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+2,end),'BFBFBF')


    start = row_size['sh_18_50'][0]
    end = row_size['sh_18_50'][1]
    images["audience_share_18_50.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('C{1}:{0}{1}'.format(next_char('b',step),start+1))

    table_border(ws,start,end,1,char_to_number(next_char('b',step)),'0')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))

    cols1 = get_column_letters_for_value(ws,start+2,end,1,char_to_number(next_char('b',step)),['FY','H1'])
    for w in cols1:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+2,end),'A4A2A2')

    cols2 = get_column_letters_for_value(ws,start+2,end,1,char_to_number(next_char('b',step)),['Q'])
    for w in cols2:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+2,end),'BFBFBF')



    start = row_size['soi'][0]
    end = row_size['soi'][1]
    images["soi_dm.png"] = "A{0}:{1}{2}".format(start,next_char('b',step*2+1),end)

    ws.merge_cells('A{0}:A{1}'.format(start,start+1))
    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step*2+1),start))

    for i,index in enumerate(range(0,step+1)):
      ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',i+index),next_char('b',i+1+index),start+1))

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step*2+1),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step*2+1),start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=char_to_number(next_char('b',step*2+1))):
        font_color = False
        if rows[0].value == 'DM Pool':
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value == 'CAS (SK, Nova Sport, Brainz)':
            font_color = True

        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if font_color:
                cell.font = Font(color='FF4149')
            if cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = '#,##0'
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0.0%'
                        except:
                            pass
                    else:
                        pass

    set_left_align(ws,'A{0}:A{1}'.format(start,end))
    

    start = row_size['cpp_4+'][0]
    end = row_size['cpp_4+'][1]
    images["average_cpp_4plus.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))

    table_border(ws,start,end,1,char_to_number(next_char('b',step)),'0.00')

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+1),'DEEAF6')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))


    start = row_size['cpp_18_50'][0]
    end = row_size['cpp_18_50'][1]
    images["average_cpp_18_50.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))

    table_border(ws,start,end,1,char_to_number(next_char('b',step)),'0.00')

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+1),'DEEAF6')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))


    start = row_size['ratio_4'][0]
    end = row_size['ratio_4'][1]
    images["power_ratio_4plus.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=char_to_number(next_char('b',step))):
      if rows[0].value in customers or rows[0].value == 'Total':
          border_type = thick
          f = True
      else:
          border_type = ColoredBorders(color='4D4D4D')
          f = False
      for cell in rows:
          cell.border = border_type
          cell.alignment = center_align
          if f:
              cell.font = bold
          if cell.value != None:
              try:
                  int(cell.value)
                  cell.number_format = '0'
              except:
                  if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0%'
                        except:
                            pass
                  else:
                      pass

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+1),'DEEAF6')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))


    start = row_size['ratio_18_50'][0]
    end = row_size['ratio_18_50'][1]
    images["power_ratio_18_50.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=char_to_number(next_char('b',step))):
      if rows[0].value in customers or rows[0].value == 'Total':
          border_type = thick
          f = True
      else:
          border_type = ColoredBorders(color='4D4D4D')
          f = False
      for cell in rows:
          cell.border = border_type
          cell.alignment = center_align
          if f:
              cell.font = bold
          if cell.value != None:
              try:
                  int(cell.value)
                  cell.number_format = '0'
              except:
                  if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0%'
                        except:
                            pass
                  else:
                      pass

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+1),'DEEAF6')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))

    return images

def cover_cas_style(ws,row_size):
    set_cell_width(ws)
    ws.column_dimensions['A'].width = 30


    start = row_size['total'][0]
    end = row_size['total'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+3),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+4),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+5),'DEEAF6')

    cas_table_border(ws,start,end,1,19,'#,##0')

    set_font_color(ws,'A{0}:S{0}'.format(start+5),'000000')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))
    
    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['sov'][0]
    end = row_size['sov'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = PatternFill("solid", fgColor="FDFF60")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+3),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+4),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+5),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+6),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+7),'DEEAF6')

    cas_table_border(ws,start,end,1,19,'0.00%')

    set_font_color(ws,'A{0}:S{0}'.format(start+5),'000000')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))
    
    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['shr'][0]
    end = row_size['shr'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws.merge_cells('A{0}:S{0}'.format(start+1))

    ws['A{0}'.format(start)].fill = pattern_fill("FF4149")
    ws['A{0}'.format(start)].font = bold
    ws['A{0}'.format(start+1)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start+1)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')

    cas_table_border(ws,start,end,1,19,'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))

    set_font_color(ws,'A{0}:S{0}'.format(start+3),'000000')
    set_font_color(ws,'A{0}:S{0}'.format(start+4),'000000')

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['18_50'][0]
    end = row_size['18_50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+3),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+4),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+5),'DEEAF6')

    cas_table_border(ws,start,end,1,19,'#,##0')

    set_font_color(ws,'A{0}:S{0}'.format(start+5),'000000')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))
    
    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['sov_18_50'][0]
    end = row_size['sov_18_50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+3),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+4),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+5),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+6),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+7),'DEEAF6')

    cas_table_border(ws,start,end,1,19,'0.00%')

    set_font_color(ws,'A{0}:S{0}'.format(start+5),'000000')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))
    
    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['shr_18_50'][0]
    end = row_size['shr_18_50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))
    ws.merge_cells('A{0}:S{0}'.format(start+1))

    ws['A{0}'.format(start)].fill = pattern_fill("FF4149")
    ws['A{0}'.format(start)].font = bold
    ws['A{0}'.format(start+1)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start+1)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')

    set_font_color(ws,'A{0}:S{0}'.format(start+3),'000000')
    set_font_color(ws,'A{0}:S{0}'.format(start+4),'000000')

    cas_table_border(ws,start,end,1,19,'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))

    row_fill(ws,'F{0}:F{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+2,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+2,end),'BFBFBF')


    start = row_size['kontrolni_duration'][0]
    end = row_size['kontrolni_duration'][1]

    ws.merge_cells('A{0}:R{0}'.format(start))
    ws.merge_cells('A{0}:R{0}'.format(start+1))

    ws['A{0}'.format(start)].fill = pattern_fill("FF4149")
    ws['A{0}'.format(start)].font = bold
    ws['A{0}'.format(start+1)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start+1)].font = bold

    row_bold_fill(ws,'A{0}:R{0}'.format(start+2),'DEEAF6')
    row_bold_fill(ws,'A{0}:R{0}'.format(start+3),'DEEAF6')
    row_bold_fill(ws,'A{0}:R{0}'.format(start+4),'DEEAF6')
    row_bold_fill(ws,'A{0}:R{0}'.format(start+5),'DEEAF6')
    row_bold_fill(ws,'A{0}:R{0}'.format(start+6),'DEEAF6')

    cas_table_border(ws,start,end,1,18,'0.00')

    set_font_color(ws,'A{0}:R{0}'.format(start+5),'000000')
    set_font_color(ws,'A{0}:R{0}'.format(start+6),'000000')
    set_font_color(ws,'A{0}:R{0}'.format(start+7),'000000')
    set_font_color(ws,'A{0}:R{0}'.format(end-1),'000000')
    set_font_color(ws,'A{0}:R{0}'.format(end),'000000')

    set_left_align(ws,'A{0}:A{1}'.format(start+2,end))
    
    row_fill(ws,'E{0}:E{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'I{0}:I{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'M{0}:M{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'Q{0}:Q{1}'.format(start+1,end),'BFBFBF')


    start = row_size['duration'][0]
    end = row_size['duration'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))

    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")
    ws['A{0}'.format(start)].font = bold

    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+2),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+3),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+4),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+5),'DEEAF6')

    cas_table_border(ws,start,end,1,19,'0.00%')

    set_font_color(ws,'A{0}:S{0}'.format(start+5),'000000')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))
    
    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


    start = row_size['soi'][0]
    end = row_size['soi'][1]

    ws.merge_cells('A{0}:AK{0}'.format(start))

    ws['A{0}'.format(start)].fill = pattern_fill("DEEAF6")
    ws['A{0}'.format(start)].font = bold

    ws.merge_cells('B{0}:C{0}'.format(start+1))
    ws.merge_cells('D{0}:E{0}'.format(start+1))
    ws.merge_cells('F{0}:G{0}'.format(start+1))
    ws.merge_cells('H{0}:I{0}'.format(start+1))
    ws.merge_cells('J{0}:K{0}'.format(start+1))
    ws.merge_cells('L{0}:M{0}'.format(start+1))
    ws.merge_cells('N{0}:O{0}'.format(start+1))
    ws.merge_cells('P{0}:Q{0}'.format(start+1))
    ws.merge_cells('R{0}:S{0}'.format(start+1))
    ws.merge_cells('T{0}:U{0}'.format(start+1))
    ws.merge_cells('V{0}:W{0}'.format(start+1))
    ws.merge_cells('X{0}:Y{0}'.format(start+1))
    ws.merge_cells('Z{0}:AA{0}'.format(start+1))
    ws.merge_cells('AB{0}:AC{0}'.format(start+1))
    ws.merge_cells('AD{0}:AE{0}'.format(start+1))
    ws.merge_cells('AF{0}:AG{0}'.format(start+1))
    ws.merge_cells('AH{0}:AI{0}'.format(start+1))
    ws.merge_cells('AJ{0}:AK{0}'.format(start+1))

    row_bold_fill(ws,'A{0}:AK{0}'.format(start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=37):
        red = False
        if rows[0].value in customers or rows[0].value == 'Total':
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value in ['CAS Others (w/o SK, Nova Sport, Brainz)','CAS "non AGB" (SK, Nova Sport, Brainz)']:
            red = True
        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if red:
                cell.font = Font(color='FF4149')
            if cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = '#,##0'
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0.0%'
                        except:
                            pass
                    else:
                        pass

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    set_font_color(ws,'A{0}:AK{0}'.format(start+3),'000000')
    set_font_color(ws,'A{0}:AK{0}'.format(start+4),'000000')

    row_fill(ws,'A{0}:AK{0}'.format(end-2),'FDFF60')
    row_fill(ws,'A{0}:AK{0}'.format(end-1),'FDFF60')
    row_fill(ws,'A{0}:AK{0}'.format(end),'FDFF60')

    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'K{0}:K{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'S{0}:S{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'Z{0}:Z{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'AA{0}:AA{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'AH{0}:AH{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'AI{0}:AI{1}'.format(start+1,end),'BFBFBF')

    start = row_size['cpp_4+'][0]
    end = row_size['cpp_4+'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))

    row_bold_fill(ws,'A{0}:S{0}'.format(start),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    cas_table_border(ws,start,end,1,19,'0.00')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')

    start = row_size['cpp_18_50'][0]
    end = row_size['cpp_18_50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))

    row_bold_fill(ws,'A{0}:S{0}'.format(start),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    cas_table_border(ws,start,end,1,19,'0.00')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')

    start = row_size['ratio'][0]
    end = row_size['ratio'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))

    row_bold_fill(ws,'A{0}:S{0}'.format(start),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=19):
        red = False
        if rows[0].value in customers or rows[0].value == 'Total':
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value in ['CAS Others (w/o SK, Nova Sport, Brainz)','CAS "non AGB" (SK, Nova Sport, Brainz)']:
            red = True
        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if red:
                cell.font = Font(color='FF4149')
            if cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = '#,##0'
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0%'
                        except:
                            pass
                    else:
                        pass

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')

    start = row_size['ratio_18_50'][0]
    end = row_size['ratio_18_50'][1]

    ws.merge_cells('A{0}:S{0}'.format(start))

    row_bold_fill(ws,'A{0}:S{0}'.format(start),'DEEAF6')
    row_bold_fill(ws,'A{0}:S{0}'.format(start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=19):
        red = False
        if rows[0].value in customers or rows[0].value == 'Total':
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value in ['CAS Others (w/o SK, Nova Sport, Brainz)','CAS "non AGB" (SK, Nova Sport, Brainz)']:
            red = True
        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if red:
                cell.font = Font(color='FF4149')
            if cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = '#,##0'
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0%'
                        except:
                            pass
                    else:
                        pass

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    row_fill(ws,'F{0}:F{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'J{0}:J{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'N{0}:N{1}'.format(start+1,end),'BFBFBF')
    row_fill(ws,'R{0}:R{1}'.format(start+1,end),'BFBFBF')


def marko_cover_cas_style(ws,quarter,row_size):
    set_cell_width(ws)
    ws.column_dimensions['A'].width = 25

    images = {}
    start = row_size['sov1'][0]
    end = row_size['sov1'][1]
    step = quarter+4
    images["sov_kupcima_dmp_mh_mp.png"] = "A{0}:{1}{2}".format(start,next_char('b',4*step+3),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',step+1),next_char('b',2*step+1),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',2*step+2),next_char('b',3*step+2),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',3*step+3),next_char('b',4*step+3),start))


    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start),'FDFF60')

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+1),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+2),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+3),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+4),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+5),'DEEAF6')


    cas_table_border(ws,start,end,1,char_to_number(next_char('b',4*step+3)),'#,##0')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))

    start = row_size['sov2'][0]
    end = row_size['sov2'][1]
    images["sov_kupcima_mediacom_om_others.png"] = "A{0}:{1}{2}".format(start,next_char('b',4*step+3),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',step+1),next_char('b',2*step+1),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',2*step+2),next_char('b',3*step+2),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',3*step+3),next_char('b',4*step+3),start))


    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start),'FDFF60')

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+1),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+2),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+3),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+4),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',4*step+3),start+5),'DEEAF6')


    cas_table_border(ws,start,end,1,char_to_number(next_char('b',4*step+3)),'#,##0')

    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))


    start = row_size['duration_1'][0]
    end = row_size['duration_1'][1]
    images["duration_dmp_mh_mp.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+1),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+2),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+3),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+4),'DEEAF6')


    cas_table_border(ws,start,end,1,char_to_number(next_char('b',step)),'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))

    cols = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',step)),['FY','ytd','Q'])
    for w in cols:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start+1,end),'BFBFBF')


    start = row_size['duration_2'][0]
    end = row_size['duration_2'][1]
    images["duration_mediacom_om_others.png"] = "A{0}:{1}{2}".format(start,next_char('b',step),end)

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+1),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+2),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+3),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',step),start+4),'DEEAF6')

    cas_table_border(ws,start,end,1,char_to_number(next_char('b',step)),'0.00%')

    set_left_align(ws,'A{0}:A{1}'.format(start,end))

    cols = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',step)),['FY','ytd','Q'])
    for w in cols:
        row_fill(ws,'{0}{1}:{0}{2}'.format(w,start,end),'BFBFBF')


    start = row_size['shr'][0]
    end = row_size['shr'][1]
    images["audience_share.png"] = "A{0}:{1}{2}".format(start,next_char('b',step*2+1),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',step+1),next_char('b',2*step+1),start))

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',2*step+1),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',2*step+1),start+1),'DEEAF6')

    ws['A{0}'.format(start)].fill = pattern_fill("FDFF60")

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=char_to_number(next_char('b',step*2+1))):
        font_color = False
        if rows[0].value in ['CAS Media','CAS Media AGB']:
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value in ['Nova Sport','Sport klub SRB']:
            font_color = True
        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if font_color:
                cell.fill = pattern_fill("FDFF60")
            if cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = '#,##0'
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0.00%'
                        except:
                            pass
                    else:
                        pass

    set_left_align(ws,'A{0}:A{1}'.format(start,end))


    start = row_size['soi'][0]
    end = row_size['soi'][1]
    images["soi_dm.png"] = "A{0}:{1}{2}".format(start,next_char('b',step*2+1),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',2*step+1),start))

    for i,index in enumerate(range(0,step+1)):
      ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',i+index),next_char('b',i+1+index),start+1))


    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',2*step+1),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',2*step+1),start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=char_to_number(next_char('b',step*2+1))):
        font_color = False
        if rows[0].value in ['CAS Media','CAS Media AGB', 'DM Pool']:
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value in ['Nova Sport','Sport klub SRB']:
            font_color = True
        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if font_color:
                cell.fill = pattern_fill("FDFF60")
            if cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = '#,##0'
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0.0%'
                        except:
                            pass
                    else:
                        pass

    set_left_align(ws,'A{0}:A{1}'.format(start,end))



    start = row_size['cpp'][0]
    end = row_size['cpp'][1]
    images["average_cpp.png"] = "A{0}:{1}{2}".format(start,next_char('b',step*2+1),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',step+1),next_char('b',2*step+1),start))

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',2*step+1),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',2*step+1),start+1),'DEEAF6')


    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=char_to_number(next_char('b',2*step+1))):
        font_color = False
        red = False
        if rows[0].value in ['CAS Media','CAS Media AGB']:
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = '0.00'
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0.00%'
                        except:
                            pass
                    else:
                        pass

    cols = get_column_letters_for_value(ws,start,end,1,char_to_number(next_char('b',2*step+1)),['FY','ytd','Q'])
    for w in cols:
        row_fill(ws,'{0}{1}:{0}{1}'.format(w,start+1),'BFBFBF')


    set_left_align(ws,'A{0}:A{1}'.format(start+1,end))


    start = row_size['ratio'][0]
    end = row_size['ratio'][1]
    images["power_ratio.png"] = "A{0}:{1}{2}".format(start,next_char('b',step*2+1),end)

    ws.merge_cells('B{1}:{0}{1}'.format(next_char('b',step),start))
    ws.merge_cells('{0}{2}:{1}{2}'.format(next_char('b',step+1),next_char('b',2*step+1),start))

    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',2*step+1),start),'DEEAF6')
    row_bold_fill(ws,'A{1}:{0}{1}'.format(next_char('b',2*step+1),start+1),'DEEAF6')

    for rows in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=char_to_number(next_char('b',step*2+1))):
        font_color = False
        red = False
        if rows[0].value in ['CAS Media','CAS Media AGB']:
            border_type = thick
            f = True
        else:
            border_type = ColoredBorders(color='4D4D4D')
            f = False
        if rows[0].value in ['Nova Sport','Sport klub SRB']:
            font_color = True
        if rows[0].value in ['CAS Others (w/o SK, Nova Sport, Brainz)','CAS "non AGB" (SK, Nova Sport, Brainz)']:
            red = True
        for cell in rows:
            cell.border = border_type
            cell.alignment = center_align
            if f:
                cell.font = bold
            if font_color:
                cell.fill = pattern_fill("FDFF60")
            if red:
                cell.font = Font(color='FF4149')
            if cell.value != None:
                try:
                    int(cell.value)
                    cell.number_format = '#,##0'
                except:
                    if '%' in cell.value:
                        try:
                            cell.value = float(cell.value.replace('%','').replace(' ',''))/100.0
                            cell.number_format = '0%'
                        except:
                            pass
                    else:
                        pass

    set_left_align(ws,'A{0}:A{1}'.format(start,end))

    return images
   
        